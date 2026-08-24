"""Evidence locator validation and reopening (spec section 12).

The verifier must be able to reopen every cited region from its locator
alone. These functions do exactly that, deterministically, against the
source root and manifest of the current run. Path containment is enforced
so locators can never escape the source root.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Sequence
from pathlib import Path

import polars as pl
import pymupdf
from docx import Document
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from data_agent.review.domain.evidence import Locator
from data_agent.review.domain.source import Source, SourceManifest, SourceType
from data_agent.review.ingestion.evidence_validator import (
    EvidenceValidationResult as LocatorValidationResult,
)
from data_agent.review.ingestion.normalizer import read_text_file

MAX_SNIPPET_CHARS = 4000
MAX_CELL_CHARS = 200

_TABULAR = frozenset({SourceType.CSV, SourceType.XLSX, SourceType.XLSM, SourceType.PARQUET})
_TEXTUAL = frozenset({SourceType.MARKDOWN, SourceType.TXT, SourceType.DOCX})


class LocatorValidationError(ValueError):
    """Raised when a locator cannot be resolved or its region cannot be reopened."""


def _resolve_source(
    locator: Locator, source_root: Path, manifest: SourceManifest
) -> tuple[Path, Source]:
    root = source_root.resolve()
    candidate = (root / locator.path).resolve()
    if candidate != root and root not in candidate.parents:
        raise LocatorValidationError(f"locator escapes the source root: {locator.path}")
    if not candidate.is_file():
        raise LocatorValidationError(f"locator file not found: {locator.path}")
    try:
        source = manifest.by_path(locator.path)
    except KeyError as exc:
        raise LocatorValidationError(
            f"locator path not in the review manifest: {locator.path}"
        ) from exc
    return candidate, source


def reopen_locator(
    uri: str,
    source_root: str | Path,
    manifest: SourceManifest,
) -> str:
    """Reopen the region cited by a locator and return it as text.

    Raises:
        LocatorValidationError: when the locator is malformed, escapes the
            source root, references an unknown file, or cites a region that
            does not exist.
    """
    from data_agent.review.ingestion.evidence_validator import EvidenceValidator

    snippet, _source = EvidenceValidator.source_backed(source_root, manifest).reopen(uri)
    return snippet


def validate_locator(
    uri: str,
    source_root: str | Path,
    manifest: SourceManifest,
) -> LocatorValidationResult:
    """Validate a locator without raising; returns a structured result."""
    from data_agent.review.ingestion.evidence_validator import EvidenceValidator

    return EvidenceValidator.source_backed(source_root, manifest).validate(uri)


def _read_region(path: Path, source: Source, locator: Locator) -> str:
    if source.source_type is SourceType.CSV:
        return _read_csv_region(path, locator)
    if source.source_type in (SourceType.XLSX, SourceType.XLSM):
        return _read_excel_region(path, source, locator)
    if source.source_type is SourceType.PARQUET:
        return _read_parquet_region(path, locator)
    if source.source_type is SourceType.PDF:
        return _read_pdf_region(path, locator)
    if source.source_type is SourceType.DOCX:
        return _read_docx_region(path, locator)
    if source.source_type in (SourceType.MARKDOWN, SourceType.TXT):
        return _read_lines_region(path, locator)
    raise LocatorValidationError(f"cannot reopen locators for {source.source_type} sources")


def _require_rows(locator: Locator, kind: str) -> tuple[int, int]:
    if locator.rows is None:
        raise LocatorValidationError(f"{kind} locators require rows=a:b")
    return locator.rows


def _require_page(locator: Locator) -> int:
    if locator.page is None:
        raise LocatorValidationError("PDF locators require page=N")
    return locator.page


def _require_lines(locator: Locator) -> tuple[int, int]:
    if locator.lines is None:
        raise LocatorValidationError("text locators require lines=a:b")
    return locator.lines


def _render_row(index: int, headers: Sequence[str], values: Sequence[object]) -> str:
    if headers:
        pairs = [
            f"{header}={_cell_text(value)}" for header, value in zip(headers, values, strict=False)
        ]
    else:
        pairs = [_cell_text(value) for value in values]
    return f"row {index}: " + ", ".join(pairs)


def _cell_text(value: object) -> str:
    text = "" if value is None else str(value)
    return text[:MAX_CELL_CHARS]


def _read_csv_region(path: Path, locator: Locator) -> str:
    start, end = _require_rows(locator, "CSV")
    text = read_text_file(path)
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if start > len(rows):
        raise LocatorValidationError(f"rows {start}:{end} out of range (file has {len(rows)} rows)")
    headers = rows[0] if rows else []
    lines = [
        _render_row(index, headers, values)
        for index, values in enumerate(rows[start - 1 : end], start=start)
    ]
    return "\n".join(lines)


def _read_excel_region(path: Path, source: Source, locator: Locator) -> str:
    start, end = _require_rows(locator, "Excel")
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        sheet = _pick_sheet(workbook, locator)
        rows = list(sheet.iter_rows(min_row=start, max_row=end, values_only=True))
        if not rows and sheet.max_row is not None and start > sheet.max_row:
            raise LocatorValidationError(
                f"rows {start}:{end} out of range (sheet {sheet.title!r} has {sheet.max_row} rows)"
            )
        header_row: tuple[object, ...] | None = next(
            (
                row
                for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True)
                if any(cell is not None for cell in row)
            ),
            None,
        )
        headers = [_cell_text(cell) for cell in (header_row or ())]
        lines = [
            _render_row(index, headers, list(values))
            for index, values in enumerate(rows, start=start)
        ]
        return "\n".join(lines)
    finally:
        workbook.close()


def _pick_sheet(workbook: Workbook, locator: Locator) -> Worksheet:
    if locator.sheet is not None:
        if locator.sheet not in workbook.sheetnames:
            raise LocatorValidationError(
                f"sheet {locator.sheet!r} not found; available: {workbook.sheetnames}"
            )
        return workbook[locator.sheet]
    if len(workbook.sheetnames) != 1:
        raise LocatorValidationError(
            f"sheet required for multi-sheet workbook: {workbook.sheetnames}"
        )
    return workbook[workbook.sheetnames[0]]


def _read_parquet_region(path: Path, locator: Locator) -> str:
    start, end = _require_rows(locator, "Parquet")
    frame = pl.read_parquet(path)
    if start > frame.height:
        raise LocatorValidationError(
            f"rows {start}:{end} out of range (file has {frame.height} rows)"
        )
    slice_ = frame.slice(start - 1, end - start + 1)
    headers = list(frame.columns)
    lines = [
        _render_row(index, headers, list(row.values()))
        for index, row in enumerate(slice_.iter_rows(named=True), start=start)
    ]
    return "\n".join(lines)


def _read_pdf_region(path: Path, locator: Locator) -> str:
    page = _require_page(locator)
    with pymupdf.open(path) as document:
        if page > document.page_count:
            raise LocatorValidationError(
                f"page {page} out of range (document has {document.page_count} pages)"
            )
        return document.load_page(page - 1).get_text()


def _read_docx_region(path: Path, locator: Locator) -> str:
    start, end = _require_lines(locator)
    document = Document(str(path))
    paragraphs = [paragraph.text for paragraph in document.paragraphs]
    if start > len(paragraphs):
        raise LocatorValidationError(
            f"lines {start}:{end} out of range (document has {len(paragraphs)} paragraphs)"
        )
    return "\n".join(
        f"line {index}: {text}"
        for index, text in enumerate(paragraphs[start - 1 : end], start=start)
    )


def _read_lines_region(path: Path, locator: Locator) -> str:
    start, end = _require_lines(locator)
    lines = read_text_file(path).splitlines()
    if start > len(lines):
        raise LocatorValidationError(
            f"lines {start}:{end} out of range (file has {len(lines)} lines)"
        )
    return "\n".join(
        f"line {index}: {text}" for index, text in enumerate(lines[start - 1 : end], start=start)
    )


def _truncate(text: str) -> str:
    if len(text) <= MAX_SNIPPET_CHARS:
        return text
    return text[:MAX_SNIPPET_CHARS] + "\n... (truncated)"
