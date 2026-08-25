"""Guarded source discovery and read tools.

The original risk-analysis application passed a ``ToolContext`` to every
callable.  The MCP server does not have a per-request Python context, so this
module turns the same boundary into a configured, read-only source root.  The
root is captured when :func:`register` is called and every user supplied path
is resolved through :func:`resolve_source_path` before it is opened.

The small, dependency-light catalogue intentionally is rebuilt for each tool
call.  That keeps metadata (size and SHA-256) honest if a source is changed and
avoids a mutable process-global manifest.  Optional readers are imported only
when their corresponding format is encountered.
"""

from __future__ import annotations

import csv
import hashlib
import os
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Annotated, Any
from urllib.parse import parse_qsl

from fastmcp import FastMCP
from fastmcp.exceptions import ToolError
from pydantic import Field

from data_agent.config import REPO_ROOT, Settings, get_settings
from data_agent.tools._safe_paths import SafePathError, resolve_relative

# These limits are deliberately conservative.  They cap model-visible output
# while still making normal source inspection useful.
DEFAULT_MAX_SOURCES = 200
DEFAULT_MAX_RESULTS = 200
DEFAULT_MAX_LINES = 200
MAX_LINE_CHARS = 500
MAX_OUTPUT_CHARS = 4_000
MAX_METADATA_FILE_BYTES = 20_000_000

_SKIP_DIRS = frozenset(
    {
        ".git",
        ".langgraph_api",
        ".mypy_cache",
        ".pytest_cache",
        ".ruff_cache",
        ".tox",
        "__pycache__",
        "build",
        "dist",
        "node_modules",
        "venv",
        ".venv",
    }
)
_SKIP_NAMES = frozenset({"thumbs.db", "desktop.ini"})
_SKIP_SUFFIXES = frozenset({".key", ".pem", ".p12", ".pfx"})

_TYPE_BY_SUFFIX = {
    ".csv": "csv",
    ".docx": "docx",
    ".json": "json",
    ".log": "txt",
    ".md": "markdown",
    ".parquet": "parquet",
    ".pdf": "pdf",
    ".txt": "txt",
    ".xlsm": "xlsm",
    ".xlsx": "xlsx",
    ".yaml": "yaml",
    ".yml": "yaml",
}
_TEXT_TYPES = frozenset({"csv", "docx", "json", "markdown", "txt", "yaml"})


class SourcePathError(ValueError):
    """Raised when a source path is not safely contained in the root."""


@dataclass(frozen=True)
class SourceMetadata:
    """Stable, serializable metadata for one file below the source root."""

    source_id: str
    path: str
    type: str
    sha256: str
    size_bytes: int
    candidate_domains: tuple[str, ...] = ()
    date_range_start: str | None = None
    date_range_end: str | None = None
    row_count: int | None = None
    sheet_names: tuple[str, ...] = ()
    column_names: tuple[str, ...] = ()
    page_count: int | None = None
    line_count: int | None = None
    parse_error: str | None = None

    def as_dict(self) -> dict[str, Any]:
        """Return JSON-compatible metadata for an MCP response."""
        return {
            "source_id": self.source_id,
            "path": self.path,
            "type": self.type,
            "sha256": self.sha256,
            "size_bytes": self.size_bytes,
            "candidate_domains": list(self.candidate_domains),
            "date_range_start": self.date_range_start,
            "date_range_end": self.date_range_end,
            "row_count": self.row_count,
            "sheet_names": list(self.sheet_names),
            "column_names": list(self.column_names),
            "page_count": self.page_count,
            "line_count": self.line_count,
            "parse_error": self.parse_error,
        }


def configured_source_root(
    root: str | os.PathLike[str] | Path | None = None,
    settings: Settings | None = None,
) -> Path:
    """Resolve the configured source root without requiring it to exist yet.

    Relative roots are anchored to the repository root, never to the process
    working directory.  ``register`` uses this function at startup, while the
    actual existence check happens at call time so the server can start before
    a mounted source directory is populated.
    """
    if root is None:
        configured = (settings or get_settings()).source_root
    else:
        configured = os.fspath(root)
    path = Path(configured)
    return (path if path.is_absolute() else REPO_ROOT / path).resolve()


def resolve_source_path(root: str | os.PathLike[str] | Path, requested: str) -> Path:
    """Resolve a source-relative path, rejecting escapes and symlinks.

    Only relative paths are accepted.  Resolving before containment checking
    closes both ``..`` traversal and symlink escape routes.  A source path must
    point to a regular file; directories are not readable tool inputs.
    """
    if not isinstance(requested, str) or not requested.strip():
        raise SourcePathError("path must not be empty")
    if "\x00" in requested:
        raise SourcePathError("path contains a NUL byte")

    root_path = Path(root).resolve()
    # Keep source-specific secret filtering separate, but reuse the common
    # containment implementation used by tabular and Python tools.
    untrusted_candidate = root_path / Path(requested.replace("\\", "/"))
    if untrusted_candidate.is_symlink():
        raise SourcePathError("symlink paths are not allowed")
    try:
        candidate = resolve_relative(root_path, requested, file_only=True)
    except (SafePathError, FileNotFoundError) as exc:
        raise SourcePathError(str(exc)) from exc
    if not _is_safe_file(candidate, root_path):
        raise SourcePathError(f"source file is not an allowed source: {requested!r}")
    return candidate


def _validate_root(root: Path) -> Path:
    root = root.resolve()
    if not root.is_dir():
        raise SourcePathError(f"configured source root does not exist: {root}")
    return root


def _is_safe_file(path: Path, root: Path) -> bool:
    if path.is_symlink() or not path.is_file():
        return False
    if (
        path.name.lower() in _SKIP_NAMES
        or path.name.startswith(".")
        or path.name.lower().startswith(".env")
    ):
        return False
    if path.suffix.lower() in _SKIP_SUFFIXES:
        return False
    try:
        relative_parts = path.relative_to(root).parts
    except ValueError:
        return False
    return not any(part in _SKIP_DIRS or part.startswith(".") for part in relative_parts[:-1])


def iter_source_files(root: str | os.PathLike[str] | Path) -> list[Path]:
    """Return regular, non-secret source files in deterministic path order."""
    root_path = _validate_root(Path(root))
    files = [path for path in root_path.rglob("*") if _is_safe_file(path, root_path)]
    return sorted(files, key=lambda path: path.relative_to(root_path).as_posix())


def file_digest(path: Path) -> tuple[str, int]:
    digest = hashlib.sha256()
    size = 0
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            size += len(chunk)
            digest.update(chunk)
    return digest.hexdigest(), size


def _read_text(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _text_lines(path: Path, source_type: str) -> list[str]:
    if source_type == "docx":
        try:
            from docx import Document
        except ImportError as exc:
            raise ValueError("DOCX reads require python-docx") from exc
        return [paragraph.text for paragraph in Document(str(path)).paragraphs]
    return _read_text(path).splitlines()


_DOMAIN_KEYWORDS = {
    "risk": "risk_metrics",
    "risk_metrics": "risk_metrics",
    "metrics": "risk_metrics",
    "var": "risk_metrics",
    "svar": "risk_metrics",
    "stress": "risk_metrics",
    "exposure": "risk_metrics",
    "sensitivity": "risk_metrics",
    "sensitivities": "risk_metrics",
    "limit": "risk_metrics",
    "limits": "risk_metrics",
    "pnl": "pnl",
    "income": "income_attribution",
    "attribution": "income_attribution",
    "controls": "post_trade_controls",
    "breach": "post_trade_controls",
    "breaches": "post_trade_controls",
    "commentary": "risk_commentary",
    "comments": "risk_commentary",
    "validation": "pnl_validation",
    "adjustment": "pnl_adjustments",
    "adjustments": "pnl_adjustments",
}


def candidate_domains_from_path(relative: str) -> tuple[str, ...]:
    found: list[str] = []
    for part in Path(relative).parts:
        key = Path(part).stem.lower()
        domain = _DOMAIN_KEYWORDS.get(key) or _DOMAIN_KEYWORDS.get(part.lower())
        if domain and domain not in found:
            found.append(domain)
    return tuple(found)


_DATE_RE = re.compile(r"(?<!\d)(\d{4}-\d{2}-\d{2})(?!\d)")


def _date_bounds(text: str) -> tuple[str | None, str | None]:
    """Extract deterministic ISO date bounds without guessing other formats."""
    values: list[date] = []
    for match in _DATE_RE.findall(text):
        try:
            parsed = date.fromisoformat(match)
        except ValueError:
            continue
        if 1990 <= parsed.year <= 2100:
            values.append(parsed)
    if not values:
        return None, None
    return min(values).isoformat(), max(values).isoformat()


def _metadata_for(path: Path, root: Path, source_id: str) -> SourceMetadata:
    relative = path.relative_to(root).as_posix()
    source_type = _TYPE_BY_SUFFIX.get(path.suffix.lower(), "unsupported")
    sha256, size_bytes = file_digest(path)
    values: dict[str, Any] = {
        "source_id": source_id,
        "path": relative,
        "type": source_type,
        "sha256": sha256,
        "size_bytes": size_bytes,
        "candidate_domains": candidate_domains_from_path(relative),
    }
    if size_bytes > MAX_METADATA_FILE_BYTES:
        values["parse_error"] = (
            f"metadata parsing skipped for files larger than {MAX_METADATA_FILE_BYTES} bytes"
        )
        return SourceMetadata(**values)

    try:
        if source_type in _TEXT_TYPES:
            text = "\n".join(_text_lines(path, source_type))
            values["line_count"] = len(text.splitlines())
            values["date_range_start"], values["date_range_end"] = _date_bounds(text)
            if source_type == "csv":
                with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                    rows = list(csv.reader(handle))
                values["row_count"] = max(len(rows) - 1, 0)
                values["column_names"] = tuple(rows[0]) if rows else ()
        elif source_type in {"xlsx", "xlsm"}:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
            try:
                names = tuple(workbook.sheetnames)
                values["sheet_names"] = names
                first_sheet = workbook[names[0]]
                values["row_count"] = first_sheet.max_row
                first = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                values["column_names"] = tuple(str(item) for item in first if item is not None)
                date_text: list[str] = []
                scanned = 0
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        date_text.extend(str(item) for item in row if item is not None)
                        scanned += len(row)
                        if scanned >= 200_000:
                            break
                    if scanned >= 200_000:
                        break
                values["date_range_start"], values["date_range_end"] = _date_bounds(
                    "\n".join(date_text)
                )
            finally:
                workbook.close()
        elif source_type == "docx":
            values["line_count"] = len(_text_lines(path, source_type))
        elif source_type == "pdf":
            import pymupdf

            with pymupdf.open(path) as document:
                values["page_count"] = document.page_count
        elif source_type == "parquet":
            try:
                from pyarrow import parquet
            except ImportError:
                import polars as pl

                frame = pl.read_parquet(path)
                values["row_count"] = frame.height
                values["column_names"] = tuple(frame.columns)
            else:
                table = parquet.read_table(path)
                values["row_count"] = table.num_rows
                values["column_names"] = tuple(table.column_names)
    except Exception as exc:  # noqa: BLE001 - metadata must still expose the file on parse failure
        values["parse_error"] = f"{type(exc).__name__}: {str(exc)[:300]}"
    return SourceMetadata(**values)


def discover_sources(
    root: str | os.PathLike[str] | Path,
    *,
    max_sources: int | None = DEFAULT_MAX_SOURCES,
) -> tuple[list[SourceMetadata], bool]:
    """Build bounded source metadata and return ``(items, truncated)``."""
    if max_sources is not None and max_sources < 1:
        raise ValueError("max_sources must be at least 1")
    root_path = _validate_root(Path(root))
    files = iter_source_files(root_path)
    limit = len(files) if max_sources is None else max_sources
    truncated = len(files) > limit
    metadata = [
        _metadata_for(path, root_path, f"SRC-{index:03d}")
        for index, path in enumerate(files[:limit], start=1)
    ]
    return metadata, truncated


def list_sources_data(
    root: str | os.PathLike[str] | Path,
    *,
    max_sources: int = DEFAULT_MAX_SOURCES,
) -> dict[str, Any]:
    """Return bounded source metadata for a configured root."""
    items, truncated = discover_sources(root, max_sources=max_sources)
    return {
        "sources": [item.as_dict() for item in items],
        "count": len(items),
        "truncated": truncated,
    }


def _readable_source_type(path: Path) -> str:
    return _TYPE_BY_SUFFIX.get(path.suffix.lower(), "unsupported")


def _bounded_range(start: int, end: int, *, limit: int = DEFAULT_MAX_LINES) -> None:
    if start < 1 or end < start:
        raise ValueError("range requires 1 <= start <= end")
    if end - start + 1 > limit:
        raise ValueError(f"range is limited to {limit} items")


def _truncate_output(text: str) -> str:
    """Keep a returned snippet at or below the model-visible character cap."""
    marker = "\n... (truncated)"
    if len(text) <= MAX_OUTPUT_CHARS:
        return text
    return text[: MAX_OUTPUT_CHARS - len(marker)] + marker


def read_lines_data(
    root: str | os.PathLike[str] | Path,
    path: str,
    start: int,
    end: int,
    *,
    max_lines: int = DEFAULT_MAX_LINES,
) -> str:
    """Read 1-based inclusive lines from a text-like source."""
    _bounded_range(start, end, limit=max_lines)
    full = resolve_source_path(root, path)
    source_type = _readable_source_type(full)
    if source_type not in _TEXT_TYPES:
        raise ValueError(f"read_lines supports text sources, not {source_type}")

    lines = _text_lines(full, source_type)
    if start > len(lines):
        raise ValueError(f"start {start} beyond file length {len(lines)}")
    if end > len(lines):
        raise ValueError(f"end {end} beyond file length {len(lines)}")
    selected = lines[start - 1 : end]
    rendered = "\n".join(
        f"line {index}: {line[:MAX_LINE_CHARS]}" for index, line in enumerate(selected, start=start)
    )
    return _truncate_output(rendered)


@dataclass(frozen=True)
class _Locator:
    path: str
    sheet: str | None = None
    page: int | None = None
    rows: tuple[int, int] | None = None
    lines: tuple[int, int] | None = None


def _positive_int(key: str, value: str) -> int:
    try:
        number = int(value)
    except ValueError as exc:
        raise ValueError(f"locator {key} must be an integer") from exc
    if number < 1:
        raise ValueError(f"locator {key} must be positive")
    return number


def _parse_range(key: str, value: str) -> tuple[int, int]:
    parts = value.split(":", 1)
    if len(parts) == 1:
        parts.append(parts[0])
    start, end = (_positive_int(key, item) for item in parts)
    if start > end:
        raise ValueError(f"locator {key} start must be <= end")
    _bounded_range(start, end)
    return start, end


def parse_source_locator(uri: str) -> _Locator:
    """Parse a strict ``source://path#sheet/page/rows/lines`` locator."""
    if not isinstance(uri, str) or not uri.startswith("source://"):
        raise ValueError("locator must start with 'source://'")
    rest = uri[len("source://") :]
    path, separator, fragment = rest.partition("#")
    if not path or path.startswith("/") or ".." in Path(path).parts:
        raise SourcePathError("locator path must stay inside the configured source root")
    values: dict[str, Any] = {"path": path}
    if separator:
        for key, value in parse_qsl(fragment, keep_blank_values=True):
            if key not in {"sheet", "page", "rows", "lines"}:
                raise ValueError(f"unknown locator fragment key {key!r}")
            if key == "sheet":
                if not value:
                    raise ValueError("locator sheet must not be empty")
                values[key] = value
            elif key == "page":
                values[key] = _positive_int(key, value)
            else:
                values[key] = _parse_range(key, value)
    selectors = [key for key in ("page", "rows", "lines") if key in values]
    if len(selectors) != 1:
        raise ValueError("locator requires exactly one of page, rows, or lines")
    return _Locator(**values)


def _render_rows(rows: list[list[Any]], start: int, headers: list[str] | None = None) -> str:
    lines = []
    for index, row in enumerate(rows, start=start):
        if headers:
            values = [
                f"{header}={'' if value is None else str(value)[:200]}"
                for header, value in zip(headers, row, strict=False)
            ]
        else:
            values = [str(value)[:200] for value in row]
        lines.append(f"row {index}: " + ", ".join(values))
    rendered = "\n".join(lines)
    return _truncate_output(rendered)


def read_document_section_data(root: str | os.PathLike[str] | Path, locator_uri: str) -> str:
    """Reopen a bounded source region from a ``source://`` locator."""
    locator = parse_source_locator(locator_uri)
    full = resolve_source_path(root, locator.path)
    source_type = _readable_source_type(full)

    if locator.page is not None:
        if source_type != "pdf":
            raise ValueError("page locators require a PDF source")
        try:
            import pymupdf
        except ImportError as exc:
            raise ValueError("PDF reads require pymupdf") from exc

        with pymupdf.open(full) as document:
            if locator.page > document.page_count:
                raise ValueError(
                    f"page {locator.page} out of range (document has {document.page_count} pages)"
                )
            text = document.load_page(locator.page - 1).get_text()
        return _truncate_output(text)

    if locator.rows is not None:
        start, end = locator.rows
        if source_type == "csv":
            with full.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
                rows = list(csv.reader(handle))
            if start > len(rows):
                raise ValueError(f"rows {start}:{end} out of range (file has {len(rows)} rows)")
            if end > len(rows):
                raise ValueError(f"rows {start}:{end} out of range (file has {len(rows)} rows)")
            return _render_rows(rows[start - 1 : end], start, rows[0] if rows else None)
        if source_type in {"xlsx", "xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ValueError("Excel reads require openpyxl") from exc

            workbook = load_workbook(full, read_only=True, data_only=True, keep_vba=False)
            try:
                if locator.sheet:
                    if locator.sheet not in workbook.sheetnames:
                        raise ValueError(f"sheet {locator.sheet!r} not found")
                    sheet = workbook[locator.sheet]
                elif len(workbook.sheetnames) == 1:
                    sheet = workbook[workbook.sheetnames[0]]
                else:
                    raise ValueError("sheet is required for a multi-sheet workbook")
                if start > sheet.max_row or end > sheet.max_row:
                    raise ValueError(
                        f"rows {start}:{end} out of range (sheet has {sheet.max_row} rows)"
                    )
                rows = [
                    list(row)
                    for row in sheet.iter_rows(min_row=start, max_row=end, values_only=True)
                ]
                header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                return _render_rows(
                    rows,
                    start,
                    ["" if value is None else str(value) for value in header],
                )
            finally:
                workbook.close()
        if source_type == "parquet":
            try:
                from pyarrow import parquet
            except ImportError:
                try:
                    import polars as pl
                except ImportError as exc:
                    raise ValueError("parquet sections require pyarrow or polars") from exc
                frame = pl.read_parquet(full)
                if start > frame.height or end > frame.height:
                    raise ValueError(
                        f"rows {start}:{end} out of range (file has {frame.height} rows)"
                    )
                records = frame.slice(start - 1, end - start + 1).to_dicts()
                return _render_rows(
                    [list(record.values()) for record in records],
                    start,
                    list(frame.columns),
                )
            table = parquet.read_table(full)
            if start > table.num_rows or end > table.num_rows:
                raise ValueError(
                    f"rows {start}:{end} out of range (file has {table.num_rows} rows)"
                )
            records = table.slice(start - 1, end - start + 1).to_pylist()
            return _render_rows(
                [list(record.values()) for record in records],
                start,
                list(table.column_names),
            )
        raise ValueError(f"row locators are not supported for {source_type} sources")

    if locator.lines is not None:
        start, end = locator.lines
        return read_lines_data(root, locator.path, start, end)

    raise ValueError("locator requires one of page, rows, or lines")


def search_text_data(
    root: str | os.PathLike[str] | Path,
    pattern: str,
    *,
    case_insensitive: bool = False,
    max_results: int = DEFAULT_MAX_RESULTS,
    path: str = ".",
) -> dict[str, Any]:
    """Regex search source text only; unlike ``grep``, never scans the workspace."""
    if not pattern.strip():
        raise ValueError("pattern must not be empty")
    if max_results < 1:
        raise ValueError("max_results must be at least 1")
    if max_results > DEFAULT_MAX_RESULTS:
        raise ValueError(f"max_results must be at most {DEFAULT_MAX_RESULTS}")
    try:
        regex = re.compile(pattern, re.IGNORECASE if case_insensitive else 0)
    except re.error as exc:
        raise ValueError(f"invalid regular expression: {exc}") from exc

    root_path = _validate_root(Path(root))
    if not isinstance(path, str) or not path.strip():
        raise SourcePathError("path must not be empty")
    if path == ".":
        target = root_path
    else:
        # A directory path is useful for narrowing an expensive source search;
        # resolve_source_path intentionally only accepts files.
        untrusted_target = root_path / Path(path.replace("\\", "/"))
        if untrusted_target.is_symlink():
            raise SourcePathError("symlink paths are not allowed")
        try:
            target = resolve_relative(root_path, path, must_exist=True, file_only=False)
        except (SafePathError, FileNotFoundError) as exc:
            raise SourcePathError(str(exc)) from exc

    candidates = (
        [target] if target.is_file() else [item for item in target.rglob("*") if item.is_file()]
    )
    matches: list[dict[str, Any]] = []
    files_scanned = 0
    truncated = False
    for full in sorted(candidates, key=lambda item: item.relative_to(root_path).as_posix()):
        if not _is_safe_file(full, root_path) or _readable_source_type(full) not in _TEXT_TYPES:
            continue
        files_scanned += 1
        if full.stat().st_size > MAX_METADATA_FILE_BYTES:
            continue
        lines = _text_lines(full, _readable_source_type(full))
        relative = full.relative_to(root_path).as_posix()
        for line_no, line in enumerate(lines, start=1):
            if regex.search(line):
                matches.append({"path": relative, "line": line_no, "text": line[:MAX_LINE_CHARS]})
                if len(matches) >= max_results:
                    truncated = True
                    break
        if truncated:
            break
    return {
        "pattern": pattern,
        "matches": matches,
        "count": len(matches),
        "truncated": truncated,
        "files_scanned": files_scanned,
    }


# Python-level facades retain the old tools' convenient direct-call shape for
# deterministic analyses and unit tests.  ``register`` below wraps these
# operations as MCP tools with a captured configured root.
def list_sources(
    root: str | os.PathLike[str] | Path,
    *,
    max_sources: int = DEFAULT_MAX_SOURCES,
) -> dict[str, Any]:
    """Direct-call facade for :func:`list_sources_data`."""
    return list_sources_data(root, max_sources=max_sources)


def search_text(
    root: str | os.PathLike[str] | Path,
    pattern: str,
    *,
    case_insensitive: bool = False,
    max_results: int = DEFAULT_MAX_RESULTS,
    path: str = ".",
) -> dict[str, Any]:
    """Direct-call facade for :func:`search_text_data`."""
    return search_text_data(
        root,
        pattern,
        case_insensitive=case_insensitive,
        max_results=max_results,
        path=path,
    )


def read_lines(
    root: str | os.PathLike[str] | Path,
    path: str,
    start: int,
    end: int,
    *,
    max_lines: int = DEFAULT_MAX_LINES,
) -> str:
    """Direct-call facade for :func:`read_lines_data`."""
    return read_lines_data(root, path, start, end, max_lines=max_lines)


def read_document_section(root: str | os.PathLike[str] | Path, locator: str) -> str:
    """Direct-call facade for :func:`read_document_section_data`."""
    return read_document_section_data(root, locator)


def register(mcp: FastMCP, root: Path | None = None, settings: Settings | None = None) -> None:
    """Attach source tools to the shared FastMCP server."""
    search_root = configured_source_root(root, settings)

    @mcp.tool
    def list_sources(
        max_sources: Annotated[int, Field(ge=1, le=DEFAULT_MAX_SOURCES)] = DEFAULT_MAX_SOURCES,
    ) -> dict[str, Any]:
        """List source files and deterministic metadata below the configured source root."""
        try:
            return list_sources_data(search_root, max_sources=max_sources)
        except (OSError, ValueError) as exc:
            raise ToolError(str(exc)) from exc

    @mcp.tool
    def search_text(
        pattern: Annotated[str, Field(description="Regular expression to find in source text.")],
        case_insensitive: Annotated[
            bool, Field(description="Ignore text case when matching.")
        ] = False,
        max_results: Annotated[int, Field(ge=1, le=DEFAULT_MAX_RESULTS)] = DEFAULT_MAX_RESULTS,
        path: Annotated[
            str, Field(description="Optional source-relative file or directory.")
        ] = ".",
    ) -> dict[str, Any]:
        """Search text-like files under the configured source root (bounded and source-only)."""
        try:
            return search_text_data(
                search_root,
                pattern,
                case_insensitive=case_insensitive,
                max_results=max_results,
                path=path,
            )
        except (OSError, ValueError) as exc:
            raise ToolError(str(exc)) from exc

    @mcp.tool
    def read_lines(
        path: Annotated[str, Field(description="Source-relative text file path.")],
        start: Annotated[int, Field(ge=1)],
        end: Annotated[int, Field(ge=1)],
        max_lines: Annotated[int, Field(ge=1, le=DEFAULT_MAX_LINES)] = DEFAULT_MAX_LINES,
    ) -> str:
        """Read bounded 1-based inclusive lines from a text-like source."""
        try:
            return read_lines_data(search_root, path, start, end, max_lines=max_lines)
        except (OSError, ValueError) as exc:
            raise ToolError(str(exc)) from exc

    @mcp.tool
    def read_document_section(
        locator: Annotated[str, Field(description="source:// locator with page, rows, or lines.")],
    ) -> str:
        """Reopen a bounded PDF page, table rows, or text lines from a source:// locator."""
        try:
            return read_document_section_data(search_root, locator)
        except (OSError, ValueError) as exc:
            raise ToolError(str(exc)) from exc


__all__ = [
    "DEFAULT_MAX_RESULTS",
    "DEFAULT_MAX_SOURCES",
    "MAX_OUTPUT_CHARS",
    "SourceMetadata",
    "SourcePathError",
    "candidate_domains_from_path",
    "configured_source_root",
    "discover_sources",
    "file_digest",
    "iter_source_files",
    "list_sources",
    "list_sources_data",
    "parse_source_locator",
    "read_document_section",
    "read_document_section_data",
    "read_lines",
    "read_lines_data",
    "register",
    "resolve_source_path",
    "search_text",
    "search_text_data",
]
