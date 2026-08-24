"""Bounded tabular inspection and read-only query tools.

The registration function binds a source root once.  MCP callers therefore
send only relative file names; no manifest or caller-controlled root is
accepted at invocation time.  CSV, TSV, Parquet, XLSX and XLSM files are
supported through Polars (with openpyxl for Excel sheets).
"""

from __future__ import annotations

import datetime as _dt
import math
import re
from pathlib import Path
from typing import Annotated, Any, Literal

import duckdb
import polars as pl
from fastmcp import FastMCP
from fastmcp.exceptions import ToolError
from pydantic import Field

from data_agent.config import REPO_ROOT, Settings, get_settings
from data_agent.tools._safe_paths import guarded_path, root_from

MAX_QUERY_ROWS = 1_000
MAX_PREVIEW_ROWS = 100
MAX_TABLE_BYTES = 100 * 1024 * 1024
MAX_REGISTERED_FILES = 500
SUPPORTED_SUFFIXES = frozenset({".csv", ".tsv", ".parquet", ".xlsx", ".xlsm"})
_AGGREGATIONS = frozenset({"sum", "mean", "count", "min", "max", "median", "std"})
_SKIP_DIRS = frozenset(
    {
        ".git",
        ".langgraph_api",
        ".mypy_cache",
        ".pytest_cache",
        ".ruff_cache",
        ".tox",
        "__pycache__",
        "build",
        "dist",
        "node_modules",
        "venv",
        ".venv",
    }
)
_SKIP_NAMES = frozenset({"thumbs.db", "desktop.ini"})
_SKIP_SUFFIXES = frozenset({".key", ".pem", ".p12", ".pfx"})


def _json_safe(value: Any) -> Any:
    """Convert common dataframe scalar values to compact JSON-safe values."""

    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return None if not math.isfinite(value) else value
    if isinstance(value, (_dt.date, _dt.datetime, _dt.time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    # Decimal, numpy scalars, and other Arrow values generally expose a useful
    # primitive through ``item``.  Fall back to text rather than leaking a
    # non-serializable object through the MCP transport.
    item = getattr(value, "item", None)
    if callable(item):
        try:
            return _json_safe(item())
        except Exception:
            pass
    return str(value)


def _frame_from_excel(path: Path, sheet: str | None) -> pl.DataFrame:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        if sheet is None:
            if len(workbook.sheetnames) != 1:
                raise ValueError(
                    "sheet is required for a workbook with multiple sheets; "
                    f"available: {workbook.sheetnames}"
                )
            sheet = workbook.sheetnames[0]
        if sheet not in workbook.sheetnames:
            raise ValueError(f"sheet {sheet!r} not found; available: {workbook.sheetnames}")
        rows: list[list[Any]] = []
        for row in workbook[sheet].iter_rows(values_only=True):
            values = list(row)
            if any(value is not None for value in values):
                rows.append(values)
            if len(rows) > 250_001:
                raise ValueError("table exceeds the 250000-row safety limit")
    finally:
        workbook.close()
    if not rows:
        return pl.DataFrame()
    headers = [
        str(value) if value is not None else f"col_{index}"
        for index, value in enumerate(rows[0])
    ]
    # Duplicate Excel headings are ambiguous and Polars rejects them.  Keep
    # names deterministic while retaining the first heading verbatim.
    seen: dict[str, int] = {}
    unique_headers: list[str] = []
    for header in headers:
        count = seen.get(header, 0)
        seen[header] = count + 1
        unique_headers.append(header if count == 0 else f"{header}_{count}")
    if len(rows) == 1:
        return pl.DataFrame({header: [] for header in unique_headers})
    data = [dict(zip(unique_headers, row, strict=False)) for row in rows[1:]]
    return pl.DataFrame(data, infer_schema_length=1_000)


def _is_safe_table_file(path: Path, root: Path) -> bool:
    """Apply source secret/hidden-path policy to a tabular file."""
    root = root.resolve()
    if path.is_symlink() or not path.is_file():
        return False
    if path.name.lower() in _SKIP_NAMES or path.name.lower().startswith(".env"):
        return False
    if path.suffix.lower() in _SKIP_SUFFIXES:
        return False
    try:
        relative_parts = path.relative_to(root).parts
    except ValueError:
        return False
    return not any(part in _SKIP_DIRS or part.startswith(".") for part in relative_parts[:-1])


def _frame_for(root: Path, path: str, sheet: str | None = None) -> pl.DataFrame:
    root = root.resolve()
    full = guarded_path(root, path)
    if not _is_safe_table_file(full, root):
        raise ToolError(f"table path is not an allowed source: {path!r}")
    if full.stat().st_size > MAX_TABLE_BYTES:
        raise ToolError(f"table exceeds the {MAX_TABLE_BYTES} byte safety limit")
    suffix = full.suffix.lower()
    try:
        if suffix == ".csv":
            return pl.read_csv(full, infer_schema_length=1_000)
        if suffix == ".tsv":
            return pl.read_csv(full, separator="\t", infer_schema_length=1_000)
        if suffix == ".parquet":
            return pl.read_parquet(full)
        if suffix in {".xlsx", ".xlsm"}:
            return _frame_from_excel(full, sheet)
    except (OSError, ValueError, pl.exceptions.PolarsError) as exc:
        raise ToolError(f"unable to read table {path!r}: {exc}") from exc
    raise ToolError("unsupported tabular file; use CSV, TSV, Parquet, XLSX, or XLSM")


def load_table(root: Path, path: str, sheet: str | None = None) -> pl.DataFrame:
    """Load one configured-root table; useful to trusted in-process analyses."""

    return _frame_for(root, path, sheet)


def inspect_table(
    root: Path,
    path: str,
    sheet: str | None = None,
    preview_rows: int = 5,
) -> dict[str, Any]:
    """Return columns, row count, and a bounded preview of one table."""

    if not 1 <= preview_rows <= MAX_PREVIEW_ROWS:
        raise ToolError(f"preview_rows must be between 1 and {MAX_PREVIEW_ROWS}")
    frame = _frame_for(root, path, sheet)
    return {
        "path": path,
        "sheet": sheet,
        "columns": list(frame.columns),
        "row_count": frame.height,
        "preview": _json_safe(frame.head(preview_rows).to_dicts()),
    }


def read_rows(
    root: Path,
    path: str,
    start: int,
    end: int,
    sheet: str | None = None,
) -> list[dict[str, Any]]:
    """Read 1-based inclusive rows, capped to the query result limit."""

    if start < 1 or end < start:
        raise ToolError("read_rows requires 1 <= start <= end")
    if end - start + 1 > MAX_QUERY_ROWS:
        raise ToolError(f"read_rows may return at most {MAX_QUERY_ROWS} rows")
    frame = _frame_for(root, path, sheet)
    if start > frame.height:
        raise ToolError(f"start {start} is beyond row count {frame.height}")
    return _json_safe(frame.slice(start - 1, end - start + 1).to_dicts())


def describe_columns(
    root: Path,
    path: str,
    sheet: str | None = None,
) -> list[dict[str, Any]]:
    """Return bounded null, distinct, type, and numeric range summaries."""

    frame = _frame_for(root, path, sheet)
    descriptions: list[dict[str, Any]] = []
    for column in frame.columns:
        series = frame[column]
        minimum: Any = None
        maximum: Any = None
        if series.dtype.is_numeric() and len(series):
            minimum = series.min()
            maximum = series.max()
        descriptions.append(
            {
                "column": column,
                "dtype": str(series.dtype),
                "null_count": series.null_count(),
                "distinct": series.n_unique(),
                "min": _json_safe(minimum),
                "max": _json_safe(maximum),
            }
        )
    return descriptions


def group_by(
    root: Path,
    path: str,
    group_columns: list[str],
    agg_column: str,
    agg: str = "sum",
    sheet: str | None = None,
) -> list[dict[str, Any]]:
    """Aggregate one column by one or more columns with a bounded result."""

    if not group_columns:
        raise ToolError("group_columns must contain at least one column")
    if agg not in _AGGREGATIONS:
        raise ToolError(f"aggregation must be one of {sorted(_AGGREGATIONS)}")
    frame = _frame_for(root, path, sheet)
    for column in [*group_columns, agg_column]:
        if column not in frame.columns:
            raise ToolError(f"unknown column {column!r}; available: {frame.columns}")
    exprs = {
        "sum": pl.col(agg_column).sum(),
        "mean": pl.col(agg_column).mean(),
        "count": pl.col(agg_column).count(),
        "min": pl.col(agg_column).min(),
        "max": pl.col(agg_column).max(),
        "median": pl.col(agg_column).median(),
        "std": pl.col(agg_column).std(),
    }
    grouped = frame.group_by(group_columns).agg(exprs[agg].alias(f"{agg}({agg_column})"))
    # Sorting makes repeated calls deterministic and keeps output stable for an
    # LLM; cap after sorting to avoid a group explosion in the response.
    return _json_safe(grouped.sort(group_columns).head(MAX_QUERY_ROWS).to_dicts())


def join_tables(
    root: Path,
    left: str,
    right: str,
    on: list[str],
    how: Literal["inner", "left", "outer", "full", "cross"] = "inner",
    left_sheet: str | None = None,
    right_sheet: str | None = None,
) -> list[dict[str, Any]]:
    """Join two configured-root tables and return at most 1000 rows."""

    if how not in {"inner", "left", "outer", "full", "cross"}:
        raise ToolError("how must be inner, left, outer, full, or cross")
    if how == "cross" and on:
        raise ToolError("cross joins must not specify join columns")
    if how != "cross" and not on:
        raise ToolError("on must contain at least one join column")
    left_frame = _frame_for(root, left, left_sheet)
    right_frame = _frame_for(root, right, right_sheet)
    for column in on:
        if column not in left_frame.columns or column not in right_frame.columns:
            raise ToolError(f"join column {column!r} must exist in both tables")
    try:
        kwargs: dict[str, Any] = {"how": "full" if how == "outer" else how}
        if how != "cross":
            kwargs["on"] = on
        joined = left_frame.join(right_frame, suffix="_right", **kwargs)
    except (ValueError, pl.exceptions.PolarsError) as exc:
        raise ToolError(f"unable to join tables: {exc}") from exc
    return _json_safe(joined.head(MAX_QUERY_ROWS).to_dicts())


def _table_name(relative: Path, sheet: str | None = None) -> str:
    stem = relative.as_posix()
    if sheet is not None:
        stem += "__" + sheet
    return "src_" + re.sub(r"[^a-zA-Z0-9_]", "_", stem).strip("_").lower()


def _iter_table_paths(root: Path) -> list[Path]:
    root = root.resolve()
    files = [
        candidate
        for candidate in root.rglob("*")
        if _is_safe_table_file(candidate, root)
        and candidate.suffix.lower() in SUPPORTED_SUFFIXES
    ]
    return sorted(files, key=lambda candidate: candidate.relative_to(root).as_posix())[
        :MAX_REGISTERED_FILES
    ]


def _duckdb_type(dtype: pl.DataType) -> str:
    """Map Polars types to stable DuckDB column types without PyArrow."""

    if dtype == pl.Boolean:
        return "BOOLEAN"
    if dtype in {pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64}:
        return "BIGINT"
    if dtype in {pl.Float32, pl.Float64}:
        return "DOUBLE"
    if dtype == pl.Date:
        return "DATE"
    if dtype == pl.Time:
        return "TIME"
    if isinstance(dtype, pl.Datetime):
        return "TIMESTAMP"
    if isinstance(dtype, pl.Decimal):
        return "DECIMAL(38, 10)"
    return "VARCHAR"


def _quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _register_frame(
    connection: duckdb.DuckDBPyConnection,
    name: str,
    frame: pl.DataFrame,
) -> None:
    """Register a frame using parameterized inserts (PyArrow is optional)."""

    if not frame.columns:
        connection.execute(f"CREATE TEMP TABLE {_quote_identifier(name)} (\"_empty\" BOOLEAN)")
        return
    columns = ", ".join(
        f"{_quote_identifier(column)} {_duckdb_type(dtype)}"
        for column, dtype in zip(frame.columns, frame.dtypes, strict=True)
    )
    table = _quote_identifier(name)
    connection.execute(f"CREATE TEMP TABLE {table} ({columns})")
    placeholders = ", ".join("?" for _ in frame.columns)
    connection.executemany(
        f"INSERT INTO {table} VALUES ({placeholders})",
        list(frame.iter_rows()),
    )


def _register_tables(root: Path, connection: duckdb.DuckDBPyConnection) -> None:
    used_names: set[str] = set()

    def register(name: str, frame: pl.DataFrame) -> None:
        original = name
        suffix = 2
        while name in used_names:
            name = f"{original}_{suffix}"
            suffix += 1
        used_names.add(name)
        _register_frame(connection, name, frame)

    for path in _iter_table_paths(root):
        relative = path.relative_to(root)
        try:
            if path.suffix.lower() in {".xlsx", ".xlsm"}:
                # Keep one-sheet workbooks convenient and expose every sheet
                # explicitly for multi-sheet workbooks.
                from openpyxl import load_workbook

                workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
                try:
                    sheets = list(workbook.sheetnames)
                finally:
                    workbook.close()
                for sheet in sheets:
                    try:
                        register(_table_name(relative, sheet), _frame_for(root, relative.as_posix(), sheet))
                    except (ToolError, ValueError, OSError):
                        continue
            else:
                register(_table_name(relative), _frame_for(root, relative.as_posix()))
        except (ToolError, ValueError, OSError):
            continue


_SQL_FORBIDDEN = re.compile(
    r"\b(?:ATTACH|CALL|COPY|CREATE|DELETE|DROP|EXPORT|IMPORT|INSERT|INSTALL|LOAD|MERGE|PRAGMA|REPLACE|RESET|SET|TRUNCATE|UPDATE|VACUUM|ALTER|BEGIN|COMMIT|ROLLBACK|TRANSACTION|READ_CSV|READ_CSV_AUTO|READ_JSON|READ_JSON_AUTO|READ_NDJSON|READ_NDJSON_AUTO|READ_PARQUET|READ_TEXT|READ_BLOB|READ_NPY|READ_XLSX|READ_XML|PARQUET_SCAN|CSV_SCAN|JSON_SCAN|GLOB|HTTPFS|DELTA_SCAN|SQLITE_SCAN|POSTGRES_SCAN|MYSQL_SCAN|NEXTVAL|SETSEED)\b",
    re.IGNORECASE,
)
_SQL_EXTERNAL_TABLE_LITERAL = re.compile(r"\b(?:FROM|JOIN)\s*['\"]", re.IGNORECASE)


def _validate_read_only_sql(sql: str) -> str:
    if not isinstance(sql, str) or not sql.strip():
        raise ToolError("sql must not be empty")
    # Strip comments and quoted literals before keyword checks. This avoids
    # rejecting harmless text values while still rejecting statement chaining.
    without_comments = re.sub(r"--[^\r\n]*|/\*.*?\*/", " ", sql, flags=re.DOTALL)
    statements = [part.strip() for part in without_comments.split(";") if part.strip()]
    if len(statements) != 1:
        raise ToolError("only one read-only SQL statement is allowed")
    statement = statements[0]
    first = re.match(r"([A-Za-z_]+)", statement)
    if not first or first.group(1).upper() not in {"SELECT", "WITH"}:
        raise ToolError("only SELECT/WITH read-only SQL is allowed")
    if _SQL_FORBIDDEN.search(statement):
        raise ToolError("mutating or external SQL statements are not allowed")
    # DuckDB treats a quoted path after FROM/JOIN as an implicit file scan,
    # which would bypass the configured table registration and root guard.
    if _SQL_EXTERNAL_TABLE_LITERAL.search(statement):
        raise ToolError("external table paths are not allowed")
    return statement


def run_duckdb_query(root: Path, sql: str, max_rows: int = MAX_QUERY_ROWS) -> list[dict[str, Any]]:
    """Run one read-only SELECT/WITH query over registered ``src_*`` tables."""

    if not 1 <= max_rows <= MAX_QUERY_ROWS:
        raise ToolError(f"max_rows must be between 1 and {MAX_QUERY_ROWS}")
    statement = _validate_read_only_sql(sql)
    try:
        with duckdb.connect(":memory:") as connection:
            _register_tables(root, connection)
            result = connection.execute(statement)
            columns = [item[0] for item in result.description or []]
            return _json_safe(
                [dict(zip(columns, row, strict=True)) for row in result.fetchmany(max_rows)]
            )
    except ToolError:
        raise
    except Exception as exc:
        raise ToolError(f"DuckDB query failed: {exc}") from exc


def register(
    mcp: FastMCP,
    root: Path | str | None = None,
    settings: Settings | None = None,
) -> None:
    """Register tabular tools against one immutable source root."""

    configured_root: Path | str | None = root
    if configured_root is None:
        configured_root = (settings or get_settings()).source_path
    bound_root = root_from(configured_root, REPO_ROOT, must_exist=False)

    @mcp.tool(name="inspect_table")
    def inspect_table_tool(
        path: Annotated[str, Field(description="Relative CSV/TSV/Parquet/Excel path.")],
        sheet: Annotated[str | None, Field(description="Excel sheet name, when needed.")] = None,
        preview_rows: Annotated[int, Field(ge=1, le=MAX_PREVIEW_ROWS)] = 5,
    ) -> dict[str, Any]:
        """Show columns, row count, and a bounded preview of a tabular source."""

        return inspect_table(bound_root, path, sheet, preview_rows)

    @mcp.tool(name="read_rows")
    def read_rows_tool(
        path: Annotated[str, Field(description="Relative tabular path.")],
        start: Annotated[int, Field(ge=1, description="First data row, 1-based.")],
        end: Annotated[int, Field(ge=1, description="Last data row, inclusive.")],
        sheet: Annotated[str | None, Field(description="Excel sheet name, when needed.")] = None,
    ) -> list[dict[str, Any]]:
        """Read a bounded inclusive range of rows from a tabular source."""

        return read_rows(bound_root, path, start, end, sheet)

    @mcp.tool(name="describe_columns")
    def describe_columns_tool(
        path: Annotated[str, Field(description="Relative tabular path.")],
        sheet: Annotated[str | None, Field(description="Excel sheet name, when needed.")] = None,
    ) -> list[dict[str, Any]]:
        """Summarize each column's type, nulls, distinct count, and numeric range."""

        return describe_columns(bound_root, path, sheet)

    @mcp.tool(name="group_by")
    def group_by_tool(
        path: Annotated[str, Field(description="Relative tabular path.")],
        group_columns: Annotated[list[str], Field(description="Columns to group by.")],
        agg_column: Annotated[str, Field(description="Column to aggregate.")],
        agg: Annotated[str, Field(description="sum, mean, count, min, max, median, or std.")] = "sum",
        sheet: Annotated[str | None, Field(description="Excel sheet name, when needed.")] = None,
    ) -> list[dict[str, Any]]:
        """Aggregate one column by grouping columns, returning at most 1000 groups."""

        return group_by(bound_root, path, group_columns, agg_column, agg, sheet)

    @mcp.tool(name="join_tables")
    def join_tables_tool(
        left: Annotated[str, Field(description="Relative path of the left table.")],
        right: Annotated[str, Field(description="Relative path of the right table.")],
        on: Annotated[list[str], Field(description="Shared join columns.")],
        how: Annotated[str, Field(description="inner, left, outer, full, or cross.")] = "inner",
        left_sheet: Annotated[str | None, Field(description="Left Excel sheet.")] = None,
        right_sheet: Annotated[str | None, Field(description="Right Excel sheet.")] = None,
    ) -> list[dict[str, Any]]:
        """Join two tabular sources and return at most 1000 rows."""

        return join_tables(bound_root, left, right, on, how, left_sheet, right_sheet)  # type: ignore[arg-type]

    @mcp.tool(name="run_duckdb_query")
    def run_duckdb_query_tool(
        sql: Annotated[str, Field(description="One read-only SELECT/WITH query over src_* tables.")],
        max_rows: Annotated[int, Field(ge=1, le=MAX_QUERY_ROWS)] = MAX_QUERY_ROWS,
    ) -> list[dict[str, Any]]:
        """Run a bounded read-only DuckDB query over configured tabular files."""

        return run_duckdb_query(bound_root, sql, max_rows)
